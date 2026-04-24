#!/usr/bin/env python3
"""IronCalc xlsx -> Formulon golden JSON importer.

Reads the xlsx fixtures vendored at
`tests/oracle/external/ironcalc/fixtures/` (originally from
https://github.com/ironcalc/IronCalc, dual MIT / Apache-2.0) and emits
one golden JSON per (xlsx, sheet) under
`tests/oracle/golden/ironcalc/`.

The emitted JSON matches Formulon's existing oracle golden schema
(`tests/oracle/README.md`): one `suite`, an `environment`, a numeric
`tolerance`, and a `cases` list where each case carries `formula`,
`setup`, and `expect` (from the xlsx's cached `<v>`).

Design notes
------------

* **Cross-sheet formulas are skipped.** Formulon's oracle runner
  operates on a single sheet per case (`wb.sheet(0)` is hard-coded in
  `tests/oracle/oracle_test.cpp`). Any formula containing a
  `SheetName!` qualifier would need multi-sheet setup that the current
  runner cannot provide, so those formulas are dropped and logged.

* **Flat output directory.** `tests/oracle/oracle_runner.cpp` uses
  `std::filesystem::directory_iterator` (non-recursive) so every golden
  JSON has to live directly under `golden/ironcalc/`. The source's
  relative path is therefore folded into the filename using `__` as a
  separator: `calc_tests/AVERAGE.xlsx` + sheet `Sheet1` becomes
  `calc_tests__AVERAGE__Sheet1.golden.json`.

* **Setup vs. formula cells.** Every non-formula cell in a sheet that
  is *not* the cell under test becomes a `setup` entry carrying its
  raw value. Other formula cells in the same sheet are emitted with
  `{kind: formula, formula: "=..."}` so the runner re-evaluates them
  through Formulon. This exercises transitive dependencies.

* **METADATA sheet.** Upstream IronCalc stashes either a bare epsilon
  (single number in `A1`) or key/value rows like `Epsilon | 5e-8` /
  `Locale | en-GB`. Both shapes are supported; absence defaults to
  `abs=5e-8, rel=0` which matches the IronCalc runner's hard-coded
  `EPS`.

* **Error names.** openpyxl returns Excel errors as strings like
  `#DIV/0!`. We keep the literal and emit `{kind: error, code: ...}`
  so the Formulon runner's `display_name_to_code` table handles them.
"""

from __future__ import annotations

import argparse
import datetime as _dt
import json
import logging
import re
import subprocess
import sys
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

import openpyxl
from openpyxl.cell.cell import Cell
from openpyxl.formula.tokenizer import Tokenizer
from openpyxl.utils import column_index_from_string, get_column_letter
from openpyxl.utils.cell import coordinate_from_string


REPO_ROOT = Path(__file__).resolve().parents[2]
DEFAULT_SOURCE_DIR = REPO_ROOT / "tests" / "oracle" / "external" / "ironcalc" / "fixtures"
DEFAULT_OUT_DIR = REPO_ROOT / "tests" / "oracle" / "golden" / "ironcalc"
DEFAULT_LOG_PATH = Path(__file__).resolve().parent / "ironcalc_import.log"

# Formulon's existing tolerance default for xlwings-generated goldens is
# 1e-12, but IronCalc's own runner uses 5e-8 so a lot of its fixtures
# commit cached values that are only accurate to ~8 digits. Use the
# upstream default unless the fixture's METADATA sheet overrides it.
DEFAULT_ABS_TOLERANCE = 5e-8
DEFAULT_REL_TOLERANCE = 0.0

# Per-sheet cap on emitted cases. Some IronCalc fixtures (notably
# ARABIC_ROMAN, which has ~20k formula cells laid out in a grid) would
# otherwise produce gigabyte-scale goldens because each case carries its
# own `setup` snapshot. Ordering is stable (sorted by row, then column)
# so truncation is deterministic.
MAX_CASES_PER_SHEET = 400

# Maximum cells expanded from a single range reference. Guards against
# whole-column / whole-row references like `A:A` that would otherwise
# balloon the setup map.
MAX_CELLS_PER_RANGE = 4096

# Known Excel error strings. Present here only so we can cheaply
# distinguish error text (`"#DIV/0!"`) from user-entered text that
# happens to start with `#`.
EXCEL_ERROR_NAMES = frozenset([
    "#NULL!",
    "#DIV/0!",
    "#VALUE!",
    "#REF!",
    "#NAME?",
    "#NUM!",
    "#N/A",
    "#SPILL!",
    "#CALC!",
    "#FIELD!",
    "#BLOCKED!",
    "#CONNECT!",
    "#EXTERNAL!",
    "#BUSY!",
    "#PYTHON!",
    "#UNKNOWN!",
    "#GETTING_DATA",  # transient Excel name; map to #N/A-ish? we skip
])

# Matches a sheet-qualified reference token: `SheetName!A1` or
# `'Sheet 1'!A1`. Used to skip formulas that would require multi-sheet
# setup. Deliberately conservative — a false positive only causes a
# case to be skipped, not miscomputed.
CROSS_SHEET_RE = re.compile(r"(?:'[^']*'|[A-Za-z_][A-Za-z0-9_\.]*)!")

# Characters we can't put in an output filename. openpyxl will happily
# accept sheet names with slashes, spaces, etc.
_FILENAME_SCRUB = re.compile(r"[^A-Za-z0-9._-]+")


logger = logging.getLogger("ironcalc_import")


# ---------------------------------------------------------------------------
# Value conversion
# ---------------------------------------------------------------------------


def _ironcalc_commit() -> str:
    """Returns the IronCalc commit recorded in README.md, or 'unknown'.

    We prefer reading the file over invoking git so the importer runs
    deterministically even without `/tmp/ironcalc` present. The commit
    SHA in README is the source of truth for this snapshot.
    """

    readme = REPO_ROOT / "tests" / "oracle" / "external" / "ironcalc" / "README.md"
    if not readme.exists():
        return "unknown"
    for line in readme.read_text(encoding="utf-8").splitlines():
        m = re.search(r"`([0-9a-f]{40})`", line)
        if m:
            return m.group(1)
    return "unknown"


def _scrub_filename(part: str) -> str:
    """Folds a path component into filesystem-safe characters."""

    cleaned = _FILENAME_SCRUB.sub("_", part).strip("._") or "sheet"
    return cleaned


def _value_to_record(value: Any, *, where: str) -> Dict[str, Any]:
    """Normalises a Python cell value into the uniform `{kind, ...}`."""

    if value is None:
        return {"kind": "blank"}
    if isinstance(value, bool):
        return {"kind": "bool", "value": value}
    if isinstance(value, (int, float)):
        f = float(value)
        return {"kind": "number", "value": f}
    if isinstance(value, str):
        # IronCalc's cached error values come through as plain strings.
        if value.startswith("#") and value in EXCEL_ERROR_NAMES:
            return {"kind": "error", "code": value}
        return {"kind": "text", "value": value}
    # Datetime / time / timedelta: translate to Excel serial so the
    # setup lands as a plain number. For now we skip rather than guess
    # since IronCalc fixtures generally use raw serials.
    raise ValueError(f"{where}: unsupported value type {type(value).__name__}")


def _expect_from_cached(cell: Cell, *, where: str) -> Optional[Dict[str, Any]]:
    """Builds an `expect` record from a cached-value cell.

    Returns None if the cached value is absent (xlsx was never opened in
    Excel, or formula is volatile). The caller must skip such cases.
    """

    value = cell.value
    if value is None:
        # data_type 'e' with value None should not happen, but guard
        # against openpyxl quirks.
        return None
    # openpyxl exposes cached errors with data_type == 'e' and value
    # like '#DIV/0!'. With data_only=True the value IS that string, so
    # _value_to_record routes it correctly already.
    return _value_to_record(value, where=where)


# ---------------------------------------------------------------------------
# METADATA sheet
# ---------------------------------------------------------------------------


def _read_metadata(wb) -> Tuple[float, float, str]:
    """Extracts (abs_tol, rel_tol, locale) from a workbook's METADATA sheet.

    Falls back to (DEFAULT_ABS_TOLERANCE, DEFAULT_REL_TOLERANCE, 'unknown')
    when the sheet is missing or unreadable. IronCalc's two observed
    shapes:

    * A single numeric in `A1` -> treated as absolute epsilon.
    * Key/value rows (column A = key, column B = value), case-insensitive
      matching against {"epsilon", "locale"}.
    """

    abs_tol = DEFAULT_ABS_TOLERANCE
    rel_tol = DEFAULT_REL_TOLERANCE
    locale = "unknown"
    sheet = None
    for name in wb.sheetnames:
        if name.lower() == "metadata":
            sheet = wb[name]
            break
    if sheet is None:
        return abs_tol, rel_tol, locale

    # Shape 1: single numeric in A1.
    a1 = sheet["A1"].value
    b1 = sheet["B1"].value
    if isinstance(a1, (int, float)) and not isinstance(a1, bool) and b1 is None:
        abs_tol = float(a1)
        return abs_tol, rel_tol, locale

    # Shape 2: key/value rows.
    for row in sheet.iter_rows(values_only=True):
        if not row or row[0] is None:
            continue
        key = str(row[0]).strip().lower()
        value = row[1] if len(row) > 1 else None
        if key == "epsilon" and isinstance(value, (int, float)):
            abs_tol = float(value)
        elif key == "locale" and isinstance(value, str):
            locale = value
    return abs_tol, rel_tol, locale


# ---------------------------------------------------------------------------
# Core per-sheet conversion
# ---------------------------------------------------------------------------


class SheetConvertStats:
    """Per-sheet counters so the summary can report skip reasons."""

    def __init__(self) -> None:
        self.formula_cells: int = 0
        self.emitted_cases: int = 0
        self.skipped_cross_sheet: int = 0
        self.skipped_no_cached: int = 0
        self.skipped_unsupported_value: int = 0
        self.skipped_over_cap: int = 0
        self.skipped_unbounded_ref: int = 0


def _expand_reference(ref: str) -> List[str]:
    """Expands an A1 range reference into a list of single-cell addresses.

    Accepts either a single cell (`A1`) or a range (`A1:B3`). Absolute
    markers (`$`) are stripped. Whole-column / whole-row references
    (`A:A`, `1:1`) are rejected with an empty list because expanding
    them would blow out the setup map.
    """

    ref = ref.replace("$", "")
    if ":" not in ref:
        # Single cell. Validate via coordinate_from_string.
        try:
            coordinate_from_string(ref)
        except Exception:
            return []
        return [ref]

    left, right = ref.split(":", 1)
    try:
        lcol, lrow = coordinate_from_string(left)
        rcol, rrow = coordinate_from_string(right)
    except Exception:
        return []
    lci = column_index_from_string(lcol)
    rci = column_index_from_string(rcol)
    cmin, cmax = sorted((lci, rci))
    rmin, rmax = sorted((lrow, rrow))
    size = (cmax - cmin + 1) * (rmax - rmin + 1)
    if size > MAX_CELLS_PER_RANGE:
        return []
    out: List[str] = []
    for r in range(rmin, rmax + 1):
        for c in range(cmin, cmax + 1):
            out.append(f"{get_column_letter(c)}{r}")
    return out


class UnboundedReference(Exception):
    """Raised when a formula contains a reference we won't expand.

    Examples: whole-column (`A:A`), whole-row (`$47:$47`), external
    workbook link. Callers translate this into a per-case skip — the
    case would otherwise need a multi-thousand-cell setup snapshot.
    """


def _referenced_cells(formula: str) -> List[str]:
    """Returns all cell addresses referenced by `formula`.

    Raises `UnboundedReference` when the formula contains a reference
    we refuse to expand. That includes whole-column / whole-row
    references and anything that fails tokenization outright.
    """

    try:
        tok = Tokenizer(formula)
    except Exception as exc:
        raise UnboundedReference(f"tokenize failed: {exc}") from exc
    refs: List[str] = []
    for item in tok.items:
        if item.type == "OPERAND" and item.subtype == "RANGE":
            expanded = _expand_reference(item.value)
            if not expanded:
                raise UnboundedReference(
                    f"cannot expand reference {item.value!r}"
                )
            refs.extend(expanded)
    return refs


def _build_setup_for(
    cell_under_test: Cell,
    non_formula_cells: Dict[str, Any],
    formula_cells: Dict[str, str],
    *,
    referenced: List[str],
) -> Dict[str, Dict[str, Any]]:
    """Assembles a `setup` map restricted to cells the formula reaches.

    `referenced` comes from `_referenced_cells` and may include formula
    cells; we walk transitively through those so the chain of deps is
    included. Raises `UnboundedReference` if any transitive formula
    cell contains an unbounded reference (whole-col/row etc.).
    """

    setup: Dict[str, Dict[str, Any]] = {}
    visited: set = set()
    stack: List[str] = list(referenced)
    formula_closure: List[str] = []
    value_closure: List[str] = []
    while stack:
        addr = stack.pop()
        if addr in visited or addr == cell_under_test.coordinate:
            continue
        visited.add(addr)
        if addr in formula_cells:
            formula_closure.append(addr)
            # _referenced_cells may raise UnboundedReference; let it
            # propagate so the caller skips this whole case.
            deeper = _referenced_cells(formula_cells[addr])
            stack.extend(deeper)
        elif addr in non_formula_cells:
            value_closure.append(addr)

    for addr in value_closure:
        value = non_formula_cells[addr]
        try:
            setup[addr] = _value_to_record(
                value, where=f"setup[{addr}]"
            )
        except ValueError:
            # Unsupported setup type (e.g. datetime) -> drop; Formulon
            # treats the address as blank, which is the safest default.
            continue
    for addr in formula_closure:
        formula = formula_cells[addr]
        if not formula.startswith("="):
            formula = "=" + formula
        setup[addr] = {"kind": "formula", "formula": formula}
    return setup


def _convert_sheet(
    sheet,
    cached_sheet,
    *,
    rel_source: str,
    xlsx_stem: str,
    abs_tol: float,
    rel_tol: float,
    locale: str,
    commit: str,
    now_iso: str,
) -> Tuple[Optional[Dict[str, Any]], SheetConvertStats]:
    """Converts one sheet of one xlsx to a golden-JSON dict.

    Returns (None, stats) when the sheet has no emittable cases.
    """

    stats = SheetConvertStats()

    # Gather values up-front so we can produce setup per formula cell.
    non_formula_cells: Dict[str, Any] = {}
    formula_cells: Dict[str, str] = {}
    for row in sheet.iter_rows():
        for cell in row:
            if cell.value is None:
                continue
            if cell.data_type == "f" and isinstance(cell.value, str):
                formula_cells[cell.coordinate] = cell.value
            else:
                non_formula_cells[cell.coordinate] = cell.value

    stats.formula_cells = len(formula_cells)

    # Deterministic iteration order: sort by (row, column) so the
    # per-sheet cap truncates consistently across runs and platforms.
    def _sort_key(addr: str) -> Tuple[int, int]:
        try:
            col, row = coordinate_from_string(addr)
        except Exception:
            return (0, 0)
        return (row, column_index_from_string(col))

    ordered_addrs = sorted(formula_cells.keys(), key=_sort_key)

    cases: List[Dict[str, Any]] = []
    for addr in ordered_addrs:
        if len(cases) >= MAX_CASES_PER_SHEET:
            # Hard cap: everything beyond this point is dropped with a
            # single summary log entry rather than one per skipped cell.
            stats.skipped_over_cap = len(ordered_addrs) - ordered_addrs.index(addr)
            logger.info(
                "cap reached: %s sheet=%r dropped=%d cells beyond #%d",
                rel_source, sheet.title, stats.skipped_over_cap,
                MAX_CASES_PER_SHEET,
            )
            break
        formula = formula_cells[addr]
        if CROSS_SHEET_RE.search(formula):
            stats.skipped_cross_sheet += 1
            logger.info(
                "skip cross-sheet: %s sheet=%r cell=%s formula=%s",
                rel_source, sheet.title, addr, formula,
            )
            continue

        cached_cell = cached_sheet[addr]
        try:
            expect = _expect_from_cached(
                cached_cell,
                where=f"{rel_source}[{sheet.title}]!{addr}",
            )
        except ValueError as exc:
            stats.skipped_unsupported_value += 1
            logger.info("skip unsupported expect: %s", exc)
            continue
        if expect is None:
            stats.skipped_no_cached += 1
            logger.info(
                "skip no-cached-value: %s sheet=%r cell=%s formula=%s",
                rel_source, sheet.title, addr, formula,
            )
            continue

        try:
            referenced = _referenced_cells(formula)
            setup = _build_setup_for(
                cached_cell, non_formula_cells, formula_cells,
                referenced=referenced,
            )
        except UnboundedReference as exc:
            stats.skipped_unbounded_ref += 1
            logger.info(
                "skip unbounded-ref: %s sheet=%r cell=%s reason=%s",
                rel_source, sheet.title, addr, exc,
            )
            continue

        formula_text = formula if formula.startswith("=") else "=" + formula
        cases.append({
            "id": addr,
            "formula": formula_text,
            "setup": setup,
            "expect": expect,
        })
        stats.emitted_cases += 1

    if not cases:
        return None, stats

    rel_slug = _scrub_filename(rel_source.replace("/", "__"))
    sheet_slug = _scrub_filename(sheet.title)
    # rel_source looks like "calc_tests/AVERAGE.xlsx"; strip suffix.
    rel_dir = _scrub_filename(
        str(Path(rel_source).parent).replace("/", "__")
    )
    if rel_dir in ("", "."):
        suite = f"ironcalc_{xlsx_stem}_{sheet_slug}"
    else:
        suite = f"ironcalc_{rel_dir}_{xlsx_stem}_{sheet_slug}"

    doc = {
        "suite": suite,
        "description": (
            f"Derived from IronCalc fixture {rel_source}, "
            f"sheet {sheet.title!r}."
        ),
        "environment": {
            "source": "ironcalc",
            "commit": commit,
            "sheet": sheet.title,
            "excel_locale": locale,
            "generated_at": now_iso,
        },
        "tolerance": {"abs": abs_tol, "rel": rel_tol},
        "cases": cases,
    }
    return doc, stats


# ---------------------------------------------------------------------------
# Filesystem walk
# ---------------------------------------------------------------------------


def _iter_xlsx(source_dir: Path) -> List[Path]:
    """All `.xlsx` files under `source_dir`, sorted, deterministic."""

    files = sorted(
        p for p in source_dir.rglob("*.xlsx")
        if p.is_file() and not p.name.startswith(".")
    )
    return files


def _out_filename(rel_source: str, xlsx_stem: str, sheet_title: str) -> str:
    """Encodes the source path + sheet into a flat filename.

    `oracle_runner.cpp` does not recurse into subdirectories, so we
    fold the relative path into the filename. The separator is `__` to
    match the existing suite naming convention and to remain visually
    distinct from single-underscore tokens.
    """

    rel_dir = Path(rel_source).parent
    sheet_slug = _scrub_filename(sheet_title)
    if str(rel_dir) in (".", ""):
        stem = f"{xlsx_stem}__{sheet_slug}"
    else:
        dir_slug = _scrub_filename(str(rel_dir).replace("/", "__"))
        stem = f"{dir_slug}__{xlsx_stem}__{sheet_slug}"
    return f"{stem}.golden.json"


# ---------------------------------------------------------------------------
# CLI entry
# ---------------------------------------------------------------------------


def _configure_logging(log_path: Path) -> None:
    log_path.parent.mkdir(parents=True, exist_ok=True)
    logger.setLevel(logging.INFO)
    for h in list(logger.handlers):
        logger.removeHandler(h)
    file_handler = logging.FileHandler(log_path, mode="w", encoding="utf-8")
    file_handler.setFormatter(logging.Formatter("%(message)s"))
    logger.addHandler(file_handler)


def main(argv: Optional[List[str]] = None) -> int:
    ap = argparse.ArgumentParser(description=__doc__)
    ap.add_argument(
        "--source", type=Path, default=DEFAULT_SOURCE_DIR,
        help=f"IronCalc fixtures root (default {DEFAULT_SOURCE_DIR})",
    )
    ap.add_argument(
        "--out", type=Path, default=DEFAULT_OUT_DIR,
        help=f"Golden JSON output directory (default {DEFAULT_OUT_DIR})",
    )
    ap.add_argument(
        "--log", type=Path, default=DEFAULT_LOG_PATH,
        help="Where to tee per-skip reasons",
    )
    ap.add_argument(
        "--strict", action="store_true",
        help="Treat any per-file exception as fatal (exit 1)",
    )
    args = ap.parse_args(argv)

    if not args.source.exists():
        print(f"ironcalc_import: source not found: {args.source}",
              file=sys.stderr)
        return 1
    args.out.mkdir(parents=True, exist_ok=True)

    _configure_logging(args.log)

    commit = _ironcalc_commit()
    now_iso = _dt.datetime.now(_dt.timezone.utc).isoformat(timespec="seconds")
    now_iso = now_iso.replace("+00:00", "Z")

    xlsx_files = _iter_xlsx(args.source)
    total_xlsx = 0
    total_goldens = 0
    total_cases = 0
    total_skip_cross_sheet = 0
    total_skip_no_cached = 0
    total_skip_unsupported = 0
    total_skip_over_cap = 0
    total_skip_unbounded = 0
    errored_files: List[Tuple[str, str]] = []

    # Purge any prior run's goldens so deletions propagate. Preserve
    # non-JSON files (README, etc.) just in case.
    for existing in args.out.iterdir() if args.out.exists() else []:
        if existing.is_file() and existing.name.endswith(".golden.json"):
            existing.unlink()

    for xlsx_path in xlsx_files:
        total_xlsx += 1
        rel_source = xlsx_path.relative_to(args.source).as_posix()
        xlsx_stem = _scrub_filename(xlsx_path.stem)
        logger.info("== %s ==", rel_source)
        try:
            wb_formula = openpyxl.load_workbook(
                xlsx_path, data_only=False, read_only=False,
            )
            wb_cached = openpyxl.load_workbook(
                xlsx_path, data_only=True, read_only=False,
            )
        except Exception as exc:
            msg = f"load failed for {rel_source}: {exc}"
            logger.info(msg)
            errored_files.append((rel_source, str(exc)))
            if args.strict:
                print(msg, file=sys.stderr)
                return 1
            continue

        try:
            abs_tol, rel_tol, locale = _read_metadata(wb_formula)
        except Exception as exc:
            logger.info("metadata read failed for %s: %s", rel_source, exc)
            abs_tol = DEFAULT_ABS_TOLERANCE
            rel_tol = DEFAULT_REL_TOLERANCE
            locale = "unknown"

        for sheet_name in wb_formula.sheetnames:
            if sheet_name.lower() == "metadata":
                continue
            try:
                sheet = wb_formula[sheet_name]
                cached_sheet = wb_cached[sheet_name]
                doc, stats = _convert_sheet(
                    sheet, cached_sheet,
                    rel_source=rel_source,
                    xlsx_stem=xlsx_stem,
                    abs_tol=abs_tol,
                    rel_tol=rel_tol,
                    locale=locale,
                    commit=commit,
                    now_iso=now_iso,
                )
            except Exception as exc:
                logger.info(
                    "convert failed for %s!%s: %s",
                    rel_source, sheet_name, exc,
                )
                errored_files.append(
                    (f"{rel_source}!{sheet_name}", str(exc))
                )
                if args.strict:
                    print(f"convert failed: {exc}", file=sys.stderr)
                    return 1
                continue

            total_skip_cross_sheet += stats.skipped_cross_sheet
            total_skip_no_cached += stats.skipped_no_cached
            total_skip_unsupported += stats.skipped_unsupported_value
            total_skip_over_cap += stats.skipped_over_cap
            total_skip_unbounded += stats.skipped_unbounded_ref

            if doc is None:
                continue

            out_name = _out_filename(rel_source, xlsx_stem, sheet_name)
            out_path = args.out / out_name
            with out_path.open("w", encoding="utf-8") as f:
                json.dump(doc, f, ensure_ascii=False, indent=2)
                f.write("\n")
            total_goldens += 1
            total_cases += len(doc["cases"])

    print(
        f"ironcalc_import: processed xlsx={total_xlsx} "
        f"goldens={total_goldens} cases={total_cases} "
        f"skipped_cross_sheet={total_skip_cross_sheet} "
        f"skipped_no_cached={total_skip_no_cached} "
        f"skipped_unsupported={total_skip_unsupported} "
        f"skipped_over_cap={total_skip_over_cap} "
        f"skipped_unbounded_ref={total_skip_unbounded} "
        f"errors={len(errored_files)}"
    )
    if errored_files:
        print("ironcalc_import: errored files (see log):",
              file=sys.stderr)
        for path, err in errored_files[:10]:
            print(f"  {path}: {err}", file=sys.stderr)
    print(f"ironcalc_import: log -> {args.log}")
    print(f"ironcalc_import: out -> {args.out}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
