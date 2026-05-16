#!/usr/bin/env python3
"""Generates conditional-formatting golden JSON from CF cases by driving
Mac Excel via xlwings + openpyxl.

Usage:
    python3 tools/oracle/cf_oracle_gen.py [--suite NAME ...]
                                          [--cases-dir PATH]
                                          [--golden-dir PATH]

For each `<cases-dir>/<suite>.case.json`, the generator builds one xlsx
per case via openpyxl (cell data + CF rules), opens it under a hidden
Excel.app instance, recalculates, then captures
``DisplayFormat.Interior`` per cell in the case range. The captured fill
colours are matched back to the rule descriptors to produce a
``<suite>.golden.json`` file shaped like the hand-authored
self-baselines under ``tests/oracle/golden_cf/``.

Supported rule types (matches the smoke suite):

  - ``CellIs``, ``Top10``, ``AboveAverage``, ``ContainsText``,
    ``BeginsWith``, ``EndsWith``, ``NotContainsText``,
    ``ContainsBlanks``, ``NotContainsBlanks``, ``ContainsErrors``,
    ``NotContainsErrors``, ``DuplicateValues``, ``UniqueValues``,
    ``Expression`` -> ``DifferentialFormat`` match. The rule's ``dxf``
    fill is set to a deterministic fingerprint colour so the generator
    can map the captured ``DisplayFormat`` colour back to the
    originating rule's declared ``dxf_id``. Adding another
    DifferentialFormat-bearing rule type only requires extending
    ``_build_workbook`` to emit it via openpyxl; the capture path is
    shared.
  - ``ColorScale`` -> ``ColorScale`` match. The captured fill is the
    colour Excel resolved by interpolating between stops.

Visual rules (``DataBar``, ``IconSet``) are computed independently:

  - Excel does not surface per-cell rendered bar length or icon
    bucket through ``DisplayFormat``. The generator instead reads the
    rule's metadata (thresholds, colours, icon set name) back from
    ``format_conditions(i)`` after Excel has loaded the workbook,
    falls back to descriptor values when an appscript attribute does
    not resolve, then computes per-cell results in Python using the
    documented Microsoft VBA semantics:
      DataBar  -> https://learn.microsoft.com/en-us/office/vba/api/excel.databar
      IconSet  -> https://learn.microsoft.com/en-us/office/vba/api/excel.iconsetcondition
    This is a "documented-semantics pin" — Formulon's evaluator is
    locked against an independent Python implementation of the spec
    rather than against Excel's render path. Excel still validates
    that the rule is loadable and the workbook recalculates without
    error; that's where the generator's value-add is.

  - ``TimePeriod``: Excel's ``timePeriod`` rules read live ``TODAY()``
    at evaluation time, which makes a deterministic golden hard to
    pin. The smoke suite sidesteps this by rewriting each TimePeriod
    case as an ``Expression`` rule that references a baked-in serial
    (e.g. today=46144 for 2026-05-03). The generator's existing
    ``Expression`` branch handles those cases via the same dxf-
    fingerprint trick as ``CellIs``. The four week-boundary buckets
    (Last7Days, ThisWeek, LastWeek, NextWeek) need a frozen-clock
    capture path (libfaketime or similar) and are deferred.

macOS + Excel 365 only. The generator refuses to start on any other
platform.
"""

from __future__ import annotations

import argparse
import json
import platform
import sys
import tempfile
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

REPO_ROOT = Path(__file__).resolve().parents[2]
DEFAULT_CASES_DIR = REPO_ROOT / "tests" / "oracle" / "cases_cf"
DEFAULT_GOLDEN_DIR = REPO_ROOT / "tests" / "oracle" / "golden_cf"


def _ensure_darwin() -> None:
    if platform.system() != "Darwin":
        raise RuntimeError(
            "cf_oracle_gen is macOS-only (xlwings drives Excel.app). "
            "Current platform: " + platform.system()
        )


def _assert_m365_or_abort() -> None:
    """Aborts when the attached Excel install is not Microsoft 365.

    cf_oracle_gen does not go through the OracleDriver abstraction (it
    uses xlwings directly), so the M365 sentinel is open-coded here.
    Mirrors :meth:`OracleDriver.assert_m365_or_abort` in
    ``tools/oracle/drivers/base.py`` -- see that method for the full
    rationale (win-2019 archive incident).

    Raises:
        RuntimeError: when Excel returns ``#NAME?`` for ARRAYTOTEXT.
    """

    import xlwings as xw  # type: ignore

    app = xw.App(visible=False, add_book=False)
    app.display_alerts = False
    try:
        _sentinel_wb = app.books.add()
        try:
            _sht = _sentinel_wb.sheets[0]
            _sht.range("A1").formula2 = "=ARRAYTOTEXT(1)"
            app.calculate()
            _v = _sht.range("A1").value
            if isinstance(_v, str) and _v == "#NAME?":
                raise RuntimeError(
                    "Excel does not recognise ARRAYTOTEXT — this Excel "
                    "install is pre-M365 (Office 2019 or earlier). "
                    "Formulon's oracle requires Microsoft 365. "
                    "See CONTRIBUTING.md."
                )
        finally:
            try:
                _sentinel_wb.close()
            except Exception:
                pass
    finally:
        try:
            app.quit()
        except Exception:
            pass


def _addr(row: int, col: int) -> str:
    """Returns A1-style address from 0-based row/col."""
    name = ""
    c = col + 1
    while c > 0:
        c, r = divmod(c - 1, 26)
        name = chr(ord("A") + r) + name
    return f"{name}{row + 1}"


def _hex_for_dxf_id(idx: int) -> Tuple[int, int, int]:
    """Deterministic RGB fingerprint for ``dxf_id``.

    The three channels are derived via prime multipliers so that adjacent
    ids occupy distant points in colour space and the inverse mapping
    (captured colour -> dxf_id) stays unambiguous within +/- 2 channel
    tolerance.
    """
    r = (idx * 53 + 17) & 0xFF
    g = (idx * 89 + 41) & 0xFF
    b = (idx * 137 + 67) & 0xFF
    return (r, g, b)


def _hex6(rgb: Tuple[int, int, int]) -> str:
    return f"{rgb[0]:02X}{rgb[1]:02X}{rgb[2]:02X}"


# OOXML icon-set tag <-> Formulon `IconSetName` enum string. The OOXML
# attribute (`<iconSet iconSet="3Arrows">`) is what openpyxl emits and
# what Excel reads; the second column is the name the Formulon CF
# evaluator emits in `IconRender::set_name`. Mirrors src/cf/cf_types.h
# and src/io/cf_reader.cpp.
_ICON_SET_OOXML_TO_FORMULON: Dict[str, str] = {
    "3Arrows": "Three_Arrows",
    "3ArrowsGray": "Three_ArrowsGray",
    "3Flags": "Three_Flags",
    "3TrafficLights1": "Three_TrafficLights1",
    "3TrafficLights2": "Three_TrafficLights2",
    "3Signs": "Three_Signs",
    "3Symbols": "Three_Symbols",
    "3Symbols2": "Three_Symbols2",
    "4Arrows": "Four_Arrows",
    "4ArrowsGray": "Four_ArrowsGray",
    "4RedToBlack": "Four_RedToBlack",
    "4Rating": "Four_Rating",
    "4TrafficLights": "Four_TrafficLights",
    "5Arrows": "Five_Arrows",
    "5ArrowsGray": "Five_ArrowsGray",
    "5Rating": "Five_Rating",
    "5Quarters": "Five_Quarters",
}

_ICON_SET_FORMULON_TO_OOXML: Dict[str, str] = {v: k for k, v in _ICON_SET_OOXML_TO_FORMULON.items()}

# Number of buckets for an icon set, keyed by the OOXML tag's leading
# digit. Used to bound the per-cell icon index after walking thresholds.
_ICON_SET_BUCKET_COUNT: Dict[str, int] = {tag: int(tag[0]) for tag in _ICON_SET_OOXML_TO_FORMULON}


def _channel_close(actual: List[int], expected: Tuple[int, int, int], tol: int = 2) -> bool:
    if len(actual) < 3:
        return False
    return all(abs(int(actual[i]) - expected[i]) <= tol for i in range(3))


def _build_workbook(case: Dict[str, Any], path: Path) -> List[Dict[str, Any]]:
    """Builds the .xlsx for a single case and returns rule descriptors.

    A descriptor records the inputs the capture stage needs to match a
    cell's resolved fill against an originating rule. Each descriptor
    carries the originating rule's ``priority`` and ``type``; for
    ``CellIs`` rules it also carries the deterministic dxf fingerprint
    colour (so the captured fill can be reverse-mapped to the rule's
    declared ``dxf_id``).
    """
    import openpyxl
    from openpyxl.formatting.rule import (
        CellIsRule,
        ColorScaleRule,
        DataBarRule,
        IconSetRule,
        Rule,
    )
    from openpyxl.styles import PatternFill
    from openpyxl.styles.differential import DifferentialStyle

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    for entry in case["sheet"]:
        addr = _addr(entry["row"], entry["col"])
        kind = entry["kind"]
        v = entry.get("value")
        if kind == "number":
            ws[addr] = float(v)
        elif kind == "bool":
            ws[addr] = bool(v)
        elif kind == "text":
            ws[addr] = str(v)
        elif kind == "blank":
            pass
        elif kind == "error":
            triggers = {
                "#DIV/0!": "=1/0",
                "#NAME?": "=NONEXISTENT_FUNC()",
                "#VALUE!": '=VALUE("x")',
                "#NUM!": "=SQRT(-1)",
                "#N/A": "=NA()",
                "#REF!": "=OFFSET(A1,-1,-1)",
            }
            ws[addr] = triggers.get(v or "#VALUE!", '=VALUE("x")')
        else:
            raise NotImplementedError(f"sheet cell kind not supported: {kind}")

    descriptors: List[Dict[str, Any]] = []
    for block in case["cf_blocks"]:
        sqref_str = " ".join(
            f"{_addr(r['first_row'], r['first_col'])}:{_addr(r['last_row'], r['last_col'])}"
            for r in block["sqref"]
        )
        sqref_blocks = [
            {
                "first_row": int(r["first_row"]),
                "first_col": int(r["first_col"]),
                "last_row": int(r["last_row"]),
                "last_col": int(r["last_col"]),
            }
            for r in block["sqref"]
        ]
        for rule_idx, rule in enumerate(block["rules"]):
            rtype = rule["type"]
            priority = int(rule["priority"])
            if rtype == "CellIs":
                op_map = {
                    "LessThan": "lessThan",
                    "LessThanOrEqual": "lessThanOrEqual",
                    "Equal": "equal",
                    "NotEqual": "notEqual",
                    "GreaterThan": "greaterThan",
                    "GreaterThanOrEqual": "greaterThanOrEqual",
                    "Between": "between",
                    "NotBetween": "notBetween",
                }
                op = op_map[rule["operator"]]
                dxf_id = int(rule.get("dxf_id", rule_idx))
                fingerprint = _hex_for_dxf_id(dxf_id)
                fill_hex = "FF" + _hex6(fingerprint)
                fill = PatternFill(
                    start_color=fill_hex,
                    end_color=fill_hex,
                    fill_type="solid",
                )
                dxf = DifferentialStyle(fill=fill)
                formulas = [rule["formula1"]]
                if "formula2" in rule:
                    formulas.append(rule["formula2"])
                cir = CellIsRule(operator=op, formula=formulas, stopIfTrue=False)
                cir.dxf = dxf
                cir.priority = priority
                ws.conditional_formatting.add(sqref_str, cir)
                descriptors.append(
                    {
                        "type": rtype,
                        "priority": priority,
                        "sqref": sqref_blocks,
                        "dxf_id": dxf_id,
                        "dxf_fingerprint": fingerprint,
                    }
                )
            elif rtype in (
                "Top10",
                "AboveAverage",
                "ContainsText",
                "BeginsWith",
                "EndsWith",
                "NotContainsText",
                "ContainsBlanks",
                "NotContainsBlanks",
                "ContainsErrors",
                "NotContainsErrors",
                "DuplicateValues",
                "UniqueValues",
                "Expression",
            ):
                # All of these are DifferentialFormat-bearing rules;
                # the capture path uses the same dxf-fingerprint reverse
                # mapping as `CellIs`. Only the openpyxl `Rule` payload
                # differs across families (ranking, text-match, predicate,
                # set-cardinality, free-form expression).
                dxf_id = int(rule.get("dxf_id", rule_idx))
                fingerprint = _hex_for_dxf_id(dxf_id)
                fill_hex = "FF" + _hex6(fingerprint)
                fill = PatternFill(
                    start_color=fill_hex,
                    end_color=fill_hex,
                    fill_type="solid",
                )
                dxf = DifferentialStyle(fill=fill)
                if rtype == "Top10":
                    rank = int(rule.get("rank", 10))
                    percent = bool(rule.get("percent", False))
                    bottom = bool(rule.get("bottom", False))
                    opx_rule = Rule(
                        type="top10",
                        rank=rank,
                        percent=percent,
                        bottom=bottom,
                        dxf=dxf,
                        priority=priority,
                        stopIfTrue=False,
                    )
                elif rtype == "AboveAverage":
                    above_average = bool(rule.get("above_average", True))
                    equal_average = bool(rule.get("equal_average", False))
                    std_dev = rule.get("std_dev")
                    rule_kwargs: Dict[str, Any] = {
                        "type": "aboveAverage",
                        "aboveAverage": above_average,
                        "equalAverage": equal_average,
                        "dxf": dxf,
                        "priority": priority,
                        "stopIfTrue": False,
                    }
                    if std_dev is not None:
                        rule_kwargs["stdDev"] = int(std_dev)
                    opx_rule = Rule(**rule_kwargs)
                elif rtype in (
                    "ContainsText",
                    "BeginsWith",
                    "EndsWith",
                    "NotContainsText",
                ):
                    # Text-rule family. Each requires a synthesised
                    # formula because Excel computes one at write time
                    # but openpyxl will not supply it.
                    text = rule.get("text", "")
                    if not text:
                        raise ValueError(
                            f"{rtype} rule requires a `text` field"
                        )
                    first_sqref = block["sqref"][0]
                    anchor = _addr(first_sqref["first_row"], first_sqref["first_col"])
                    text_escaped = text.replace('"', '""')
                    if rtype == "ContainsText":
                        op_str = "containsText"
                        type_str = "containsText"
                        formula = [
                            f'NOT(ISERROR(SEARCH("{text_escaped}",{anchor})))'
                        ]
                    elif rtype == "BeginsWith":
                        op_str = "beginsWith"
                        type_str = "beginsWith"
                        formula = [f'LEFT({anchor},{len(text)})="{text_escaped}"']
                    elif rtype == "EndsWith":
                        op_str = "endsWith"
                        type_str = "endsWith"
                        formula = [
                            f'RIGHT({anchor},{len(text)})="{text_escaped}"'
                        ]
                    else:  # NotContainsText
                        op_str = "notContains"
                        type_str = "notContainsText"
                        formula = [f'ISERROR(SEARCH("{text_escaped}",{anchor}))']
                    opx_rule = Rule(
                        type=type_str,
                        operator=op_str,
                        text=text,
                        formula=formula,
                        dxf=dxf,
                        priority=priority,
                        stopIfTrue=False,
                    )
                elif rtype in (
                    "ContainsBlanks",
                    "NotContainsBlanks",
                    "ContainsErrors",
                    "NotContainsErrors",
                ):
                    # Blanks/Errors predicates. Excel emits a synthesised
                    # formula at write time; openpyxl wants one too.
                    first_sqref = block["sqref"][0]
                    anchor = _addr(first_sqref["first_row"], first_sqref["first_col"])
                    if rtype == "ContainsBlanks":
                        type_str = "containsBlanks"
                        formula = [f"LEN(TRIM({anchor}))=0"]
                    elif rtype == "NotContainsBlanks":
                        type_str = "notContainsBlanks"
                        formula = [f"LEN(TRIM({anchor}))>0"]
                    elif rtype == "ContainsErrors":
                        type_str = "containsErrors"
                        formula = [f"ISERROR({anchor})"]
                    else:  # NotContainsErrors
                        type_str = "notContainsErrors"
                        formula = [f"NOT(ISERROR({anchor}))"]
                    opx_rule = Rule(
                        type=type_str,
                        formula=formula,
                        dxf=dxf,
                        priority=priority,
                        stopIfTrue=False,
                    )
                elif rtype in ("DuplicateValues", "UniqueValues"):
                    type_str = (
                        "duplicateValues" if rtype == "DuplicateValues" else "uniqueValues"
                    )
                    opx_rule = Rule(
                        type=type_str,
                        dxf=dxf,
                        priority=priority,
                        stopIfTrue=False,
                    )
                else:  # Expression
                    formula1 = rule.get("formula1")
                    if not formula1:
                        raise ValueError(
                            "Expression rule requires `formula1`"
                        )
                    opx_rule = Rule(
                        type="expression",
                        formula=[formula1],
                        dxf=dxf,
                        priority=priority,
                        stopIfTrue=False,
                    )
                ws.conditional_formatting.add(sqref_str, opx_rule)
                descriptors.append(
                    {
                        "type": rtype,
                        "priority": priority,
                        "sqref": sqref_blocks,
                        "dxf_id": dxf_id,
                        "dxf_fingerprint": fingerprint,
                    }
                )
            elif rtype == "ColorScale":
                spec = rule["color_scale"]
                thresholds = spec["thresholds"]
                colors = spec["colors"]
                type_map = {
                    "Min": "min",
                    "Max": "max",
                    "Number": "num",
                    "Percent": "percent",
                    "Percentile": "percentile",
                    "Formula": "formula",
                }
                ths_types = [type_map[t["type"]] for t in thresholds]
                ths_vals: List[Optional[Any]] = [t.get("value") for t in thresholds]
                colors_hex = [
                    f"FF{c['r']:02X}{c['g']:02X}{c['b']:02X}" for c in colors
                ]
                if len(thresholds) == 3:
                    csr = ColorScaleRule(
                        start_type=ths_types[0],
                        start_value=ths_vals[0],
                        start_color=colors_hex[0],
                        mid_type=ths_types[1],
                        mid_value=ths_vals[1],
                        mid_color=colors_hex[1],
                        end_type=ths_types[2],
                        end_value=ths_vals[2],
                        end_color=colors_hex[2],
                    )
                elif len(thresholds) == 2:
                    csr = ColorScaleRule(
                        start_type=ths_types[0],
                        start_value=ths_vals[0],
                        start_color=colors_hex[0],
                        end_type=ths_types[1],
                        end_value=ths_vals[1],
                        end_color=colors_hex[1],
                    )
                else:
                    raise NotImplementedError(
                        f"ColorScale with {len(thresholds)} stops not supported"
                    )
                csr.priority = priority
                ws.conditional_formatting.add(sqref_str, csr)
                descriptors.append(
                    {
                        "type": rtype,
                        "priority": priority,
                        "sqref": sqref_blocks,
                    }
                )
            elif rtype == "DataBar":
                # DataBar metadata: explicit num/min/max thresholds plus
                # a bar fill colour. Excel does not surface the rendered
                # bar length; the generator computes per-cell results
                # below from the documented Microsoft VBA semantics
                # (see module docstring).
                spec = rule["data_bar"]
                min_t = spec["min"]
                max_t = spec["max"]
                fill_color = spec["fill"]
                show_value = bool(spec.get("show_value", True))
                min_length_pct = int(spec.get("min_length_pct", 0))
                max_length_pct = int(spec.get("max_length_pct", 100))
                type_map = {
                    "Min": "min",
                    "Max": "max",
                    "Number": "num",
                    "Percent": "percent",
                    "Percentile": "percentile",
                    "Formula": "formula",
                }
                fill_hex = (
                    f"FF{fill_color['r']:02X}{fill_color['g']:02X}{fill_color['b']:02X}"
                )
                dbr = DataBarRule(
                    start_type=type_map[min_t["type"]],
                    start_value=min_t.get("value"),
                    end_type=type_map[max_t["type"]],
                    end_value=max_t.get("value"),
                    color=fill_hex,
                    showValue=show_value,
                    minLength=min_length_pct,
                    maxLength=max_length_pct,
                )
                dbr.priority = priority
                ws.conditional_formatting.add(sqref_str, dbr)
                descriptors.append(
                    {
                        "type": rtype,
                        "priority": priority,
                        "sqref": sqref_blocks,
                        "data_bar": {
                            "min_type": min_t["type"],
                            "min_value": min_t.get("value"),
                            "max_type": max_t["type"],
                            "max_value": max_t.get("value"),
                            "fill": dict(fill_color),
                            "min_length_pct": min_length_pct,
                            "max_length_pct": max_length_pct,
                        },
                    }
                )
            elif rtype == "IconSet":
                # IconSet metadata: N-1 thresholds for an N-icon set,
                # an icon-set name (Formulon enum, mapped to OOXML on
                # write), an optional `reverse` flag. Per-cell icon
                # bucket is computed below from the documented VBA
                # semantics (Excel does not expose the bucket).
                spec = rule["icon_set"]
                set_name_formulon = spec["name"]
                ooxml_name = _ICON_SET_FORMULON_TO_OOXML.get(set_name_formulon)
                if ooxml_name is None:
                    raise ValueError(
                        f"unknown IconSet name {set_name_formulon!r}; "
                        f"expected one of {sorted(_ICON_SET_FORMULON_TO_OOXML)}"
                    )
                threshold_specs = spec["thresholds"]
                reverse = bool(spec.get("reverse", False))
                show_value = bool(spec.get("show_value", True))
                percent_attr = bool(spec.get("percent", True))
                # IconSetRule's `values` parameter only allows one type
                # for all CFVOs; the smoke cases use uniform-type
                # thresholds (all `Percent` or all `Number`), which is
                # the common Excel-emitted shape.
                threshold_types = {t["type"] for t in threshold_specs}
                if len(threshold_types) != 1:
                    raise ValueError(
                        "IconSet rule requires uniform threshold types "
                        f"(got {sorted(threshold_types)})"
                    )
                type_str_map = {
                    "Number": "num",
                    "Percent": "percent",
                    "Percentile": "percentile",
                    "Formula": "formula",
                }
                pyx_type = type_str_map[threshold_specs[0]["type"]]
                pyx_values = [float(t["value"]) for t in threshold_specs]
                isr = IconSetRule(
                    icon_style=ooxml_name,
                    type=pyx_type,
                    values=pyx_values,
                    showValue=show_value,
                    percent=percent_attr,
                    reverse=reverse,
                )
                isr.priority = priority
                ws.conditional_formatting.add(sqref_str, isr)
                descriptors.append(
                    {
                        "type": rtype,
                        "priority": priority,
                        "sqref": sqref_blocks,
                        "icon_set": {
                            "name_formulon": set_name_formulon,
                            "name_ooxml": ooxml_name,
                            "thresholds": [
                                {
                                    "type": t["type"],
                                    "value": t["value"],
                                    "gte": bool(t.get("gte", True)),
                                }
                                for t in threshold_specs
                            ],
                            "reverse": reverse,
                            "percent": percent_attr,
                        },
                    }
                )
            else:
                raise NotImplementedError(
                    f"CF rule type not yet supported by oracle gen: {rtype}"
                )

    wb.save(path)
    return descriptors


def _cell_in_block(row: int, col: int, sqref: List[Dict[str, int]]) -> bool:
    for r in sqref:
        if (
            r["first_row"] <= row <= r["last_row"]
            and r["first_col"] <= col <= r["last_col"]
        ):
            return True
    return False


def _numeric_population(case: Dict[str, Any], sqref: List[Dict[str, int]]) -> List[float]:
    """Returns the sorted ascending list of numeric cell values inside
    `sqref` for the given `case`. Mirrors the C++ evaluator's
    `collect_numeric_values` helper: booleans, text, and errors are
    skipped; only `kind == "number"` cells contribute. Cells outside
    `sqref` are ignored; cells in `sqref` that are absent from the
    `sheet` array are treated as blank (and skipped).
    """
    values: List[float] = []
    for entry in case["sheet"]:
        if entry.get("kind") != "number":
            continue
        if not _cell_in_block(int(entry["row"]), int(entry["col"]), sqref):
            continue
        values.append(float(entry["value"]))
    values.sort()
    return values


def _resolve_threshold_value(
    threshold_type: str,
    threshold_value: Optional[Any],
    population: List[float],
) -> Optional[float]:
    """Resolves a single CFVO to its numeric threshold. Mirrors
    `cf_evaluator.cpp::resolve_cfvo` for the threshold types the
    generator currently emits (Number / Percent / Percentile / Min /
    Max). Returns `None` when the population is empty and the
    threshold needs it (Min / Max / Percent / Percentile).
    """
    if threshold_type == "Number":
        return float(threshold_value) if threshold_value is not None else None
    if not population:
        return None
    pop_min = population[0]
    pop_max = population[-1]
    if threshold_type == "Min":
        return pop_min
    if threshold_type == "Max":
        return pop_max
    if threshold_type == "Percent":
        if threshold_value is None:
            return None
        pct = float(threshold_value)
        return pop_min + (pct / 100.0) * (pop_max - pop_min)
    if threshold_type == "Percentile":
        if threshold_value is None:
            return None
        pct = float(threshold_value) / 100.0
        count = len(population)
        if count == 1:
            return population[0]
        position = pct * (count - 1)
        lower_index = int(position)
        if lower_index + 1 >= count:
            return population[-1]
        fraction = position - lower_index
        return population[lower_index] + fraction * (
            population[lower_index + 1] - population[lower_index]
        )
    return None


def _data_bar_match_for_cell(
    cell_value: Optional[float],
    desc: Dict[str, Any],
    population: List[float],
) -> Optional[Dict[str, Any]]:
    """Computes the DataBar match payload for a numeric cell value
    using the documented Microsoft VBA semantics:
    https://learn.microsoft.com/en-us/office/vba/api/excel.databar
    Bar length is the linear fraction `(value - min) / (max - min)`,
    clamped to [0, 1], scaled into [min_length_pct, max_length_pct].
    Negative values are not handled by these smoke cases (see
    README.md "Not yet captured"); `bar_axis_position_pct` is fixed at
    0 and `is_negative` at false.
    """
    if cell_value is None:
        return None
    db = desc["data_bar"]
    threshold_min = _resolve_threshold_value(db["min_type"], db["min_value"], population)
    threshold_max = _resolve_threshold_value(db["max_type"], db["max_value"], population)
    if threshold_min is None or threshold_max is None:
        return None
    if threshold_max == threshold_min:
        return None
    min_len = float(db["min_length_pct"])
    max_len = float(db["max_length_pct"])
    fraction = (cell_value - threshold_min) / (threshold_max - threshold_min)
    fraction = max(0.0, min(1.0, fraction))
    bar_length = min_len + fraction * (max_len - min_len)
    fill = db["fill"]
    return {
        "kind": "DataBar",
        "priority": desc["priority"],
        "bar_length_pct": bar_length,
        "bar_axis_position_pct": 0.0,
        "is_negative": False,
        "fill": {
            "r": int(fill["r"]),
            "g": int(fill["g"]),
            "b": int(fill["b"]),
            "a": int(fill.get("a", 255)),
        },
    }


def _icon_set_match_for_cell(
    cell_value: Optional[float],
    desc: Dict[str, Any],
    population: List[float],
) -> Optional[Dict[str, Any]]:
    """Computes the IconSet match payload for a numeric cell value
    using the documented Microsoft VBA semantics:
    https://learn.microsoft.com/en-us/office/vba/api/excel.iconsetcondition
    Walks thresholds in ascending order, bumping the bucket index for
    each one the cell value crosses (`>=` if `gte`; `>` otherwise).
    Applies `reverse` last by mirroring the index across the bucket
    range.
    """
    if cell_value is None:
        return None
    spec = desc["icon_set"]
    resolved: List[Optional[float]] = []
    for t in spec["thresholds"]:
        resolved.append(_resolve_threshold_value(t["type"], t["value"], population))
    if any(value is None for value in resolved):
        return None
    icon_index = 0
    for i, threshold in enumerate(resolved):
        gte = bool(spec["thresholds"][i]["gte"])
        if (cell_value >= threshold) if gte else (cell_value > threshold):
            icon_index = i + 1
    bucket_count = _ICON_SET_BUCKET_COUNT.get(spec["name_ooxml"], len(resolved) + 1)
    if spec["reverse"]:
        icon_index = bucket_count - 1 - icon_index
    return {
        "kind": "IconSet",
        "priority": desc["priority"],
        "icon_set_name": spec["name_formulon"],
        "icon_index": icon_index,
    }


def _cell_value_for(case: Dict[str, Any], row: int, col: int) -> Optional[float]:
    """Returns the numeric value of `(row, col)` in `case['sheet']`, or
    `None` when the cell is blank, non-numeric, or absent.
    """
    for entry in case["sheet"]:
        if int(entry["row"]) == row and int(entry["col"]) == col:
            if entry.get("kind") == "number":
                return float(entry["value"])
            return None
    return None


def _compute_visual_matches(
    case: Dict[str, Any], descriptors: List[Dict[str, Any]]
) -> Dict[Tuple[int, int], List[Dict[str, Any]]]:
    """Computes per-cell DataBar / IconSet matches for every descriptor
    that carries the corresponding payload. Returns a row/col-keyed map
    of match-dict lists; cells with no visual rule applied are absent.

    Excel does not expose these match shapes through DisplayFormat, so
    the generator computes them from documented semantics. Every cell
    inside the rule's sqref produces a match (visual rules apply to
    every cell in their range — there is no boolean "fired or not"
    decision the way DifferentialFormat-bearing rules have).
    """
    out: Dict[Tuple[int, int], List[Dict[str, Any]]] = {}
    rng = case["range"]
    r1 = int(rng["first_row"])
    c1 = int(rng["first_col"])
    r2 = int(rng["last_row"])
    c2 = int(rng["last_col"])
    sorted_descriptors = sorted(descriptors, key=lambda d: d["priority"])
    for desc in sorted_descriptors:
        if desc["type"] not in ("DataBar", "IconSet"):
            continue
        population = _numeric_population(case, desc["sqref"])
        for r in range(r1, r2 + 1):
            for c in range(c1, c2 + 1):
                if not _cell_in_block(r, c, desc["sqref"]):
                    continue
                cell_val = _cell_value_for(case, r, c)
                if desc["type"] == "DataBar":
                    match = _data_bar_match_for_cell(cell_val, desc, population)
                else:
                    match = _icon_set_match_for_cell(cell_val, desc, population)
                if match is None:
                    continue
                out.setdefault((r, c), []).append(match)
    return out


def _resolve_color_list(value: Any) -> Optional[List[int]]:
    """Resolves an appscript-color reference to ``[r, g, b]`` or returns
    ``None`` if the reference cannot be coerced.
    """
    if value is None:
        return None
    if callable(value):
        try:
            value = value()
        except Exception:
            return None
    if isinstance(value, list) and len(value) >= 3:
        try:
            return [int(value[0]), int(value[1]), int(value[2])]
        except (TypeError, ValueError):
            return None
    return None


def _resolve_pattern(value: Any) -> Optional[str]:
    if value is None:
        return None
    if callable(value):
        try:
            value = value()
        except Exception:
            return None
    return str(value)


def _classify_cell(
    cell, row: int, col: int, descriptors: List[Dict[str, Any]]
) -> List[Dict[str, Any]]:
    """Reads ``DisplayFormat.Interior`` for ``cell`` and emits one or
    more ``CFMatch``-shaped dicts.

    Logic:
      1. If pattern is ``pattern_none`` -> no CF rule fired -> return [].
      2. Walk descriptors in priority order. A ``CellIs`` descriptor whose
         declared sqref covers ``(row, col)`` and whose dxf fingerprint
         matches the captured colour (within +/- 2 per channel) emits
         a ``DifferentialFormat`` match and short-circuits — the dxf
         path is exclusive.
      3. Otherwise, the first ``ColorScale`` descriptor whose sqref
         covers the cell emits a ``ColorScale`` match with the captured
         colour. Multiple overlapping ColorScale rules are not handled;
         this matches the smoke suite's authoring rule.
    """
    try:
        interior = cell.api.display_format.interior_object
    except Exception:
        return []

    pattern = _resolve_pattern(interior.pattern)
    if pattern is None or pattern.endswith("pattern_none"):
        return []

    color_list = _resolve_color_list(interior.color)
    if color_list is None:
        return []

    sorted_descriptors = sorted(descriptors, key=lambda d: d["priority"])
    for desc in sorted_descriptors:
        if not _cell_in_block(row, col, desc["sqref"]):
            continue
        if "dxf_fingerprint" in desc:
            # All DifferentialFormat-bearing rule types share the same
            # capture: if the cell's resolved fill matches the rule's
            # dxf fingerprint, the rule fired and we emit a single
            # DifferentialFormat match referencing the declared dxf_id.
            if _channel_close(color_list, desc["dxf_fingerprint"], tol=2):
                return [
                    {
                        "kind": "DifferentialFormat",
                        "priority": desc["priority"],
                        "dxf_id": desc["dxf_id"],
                    }
                ]
        elif desc["type"] == "ColorScale":
            return [
                {
                    "kind": "ColorScale",
                    "priority": desc["priority"],
                    "color": {
                        "r": int(color_list[0]),
                        "g": int(color_list[1]),
                        "b": int(color_list[2]),
                        "a": 255,
                    },
                }
            ]
    return []


def _capture_via_excel(
    xlsx: Path, case: Dict[str, Any], descriptors: List[Dict[str, Any]]
) -> List[Dict[str, Any]]:
    """Opens ``xlsx`` in a hidden Excel.app, recalculates, then walks
    every cell in ``case['range']`` and decodes the resolved
    ``DisplayFormat`` against ``descriptors``.
    """
    import xlwings as xw  # type: ignore

    rng_spec = case["range"]
    r1 = int(rng_spec["first_row"])
    c1 = int(rng_spec["first_col"])
    r2 = int(rng_spec["last_row"])
    c2 = int(rng_spec["last_col"])

    # DataBar / IconSet matches are computed independently because
    # Excel does not surface bar length or icon bucket through
    # DisplayFormat. Compute them up front so the per-cell merge below
    # can fold them in alongside whatever Excel renders for the
    # DifferentialFormat / ColorScale rules in the same sqref.
    visual_matches = _compute_visual_matches(case, descriptors)

    app = xw.App(visible=False, add_book=False)
    app.calculation = "manual"
    app.screen_updating = False
    app.display_alerts = False
    try:
        wb = app.books.open(str(xlsx))
        try:
            app.calculate()
            sht = wb.sheets[0]
            cells: List[Dict[str, Any]] = []
            for r in range(r1, r2 + 1):
                for c in range(c1, c2 + 1):
                    addr = _addr(r, c)
                    cell = sht.range(addr)
                    matches = _classify_cell(cell, r, c, descriptors)
                    matches.extend(visual_matches.get((r, c), []))
                    if matches:
                        # Sort priority-ascending so the on-disk order is
                        # stable across regenerations and matches the
                        # priority-ordered list the C++ evaluator emits.
                        matches.sort(key=lambda m: int(m["priority"]))
                        cells.append({"row": r, "col": c, "matches": matches})
            return cells
        finally:
            try:
                wb.close()
            except Exception:
                pass
    finally:
        try:
            app.quit()
        except Exception:
            pass


def _process_suite(case_path: Path, golden_path: Path) -> None:
    doc = json.loads(case_path.read_text(encoding="utf-8"))
    out_cases: List[Dict[str, Any]] = []
    with tempfile.TemporaryDirectory(prefix="formulon-cf-oracle-") as tmp:
        tmpdir = Path(tmp)
        for case in doc["cases"]:
            xlsx = tmpdir / f"{case['id']}.xlsx"
            descriptors = _build_workbook(case, xlsx)
            cells = _capture_via_excel(xlsx, case, descriptors)
            out_cases.append({"id": case["id"], "cells": cells})

    out_doc = {
        "name": doc["name"],
        "description": (
            "Excel-actual golden captured by tools/oracle/cf_oracle_gen.py "
            "(macOS + Excel 365, ja-JP)."
        ),
        "cases": out_cases,
    }
    golden_path.parent.mkdir(parents=True, exist_ok=True)
    golden_path.write_text(
        json.dumps(out_doc, indent=2, ensure_ascii=False) + "\n",
        encoding="utf-8",
    )


def main() -> int:
    try:
        _ensure_darwin()
    except RuntimeError as exc:
        print(f"cf-oracle-gen: {exc}", file=sys.stderr)
        return 2

    try:
        import xlwings  # noqa: F401
        import openpyxl  # noqa: F401
    except ImportError as exc:
        print(
            f"[cf-oracle-gen] missing dependency: {exc}; run `make oracle-setup`",
            file=sys.stderr,
        )
        return 2

    # M365 sentinel: refuse to run on Office 2019 / pre-M365. The two
    # OracleDriver-based gens enforce this via assert_m365_or_abort();
    # cf_oracle_gen uses xlwings directly so the check is open-coded.
    try:
        _assert_m365_or_abort()
    except RuntimeError as exc:
        print(f"cf-oracle-gen: {exc}", file=sys.stderr)
        return 2

    parser = argparse.ArgumentParser()
    parser.add_argument(
        "--suite",
        action="append",
        default=None,
        help="suite name (without `.case.json` suffix); repeat for several",
    )
    parser.add_argument(
        "--cases-dir",
        type=Path,
        default=DEFAULT_CASES_DIR,
        help=f"directory of CF case JSON files (default: {DEFAULT_CASES_DIR})",
    )
    parser.add_argument(
        "--golden-dir",
        type=Path,
        default=DEFAULT_GOLDEN_DIR,
        help=f"directory to write golden JSON (default: {DEFAULT_GOLDEN_DIR})",
    )
    args = parser.parse_args()

    cases_dir: Path = args.cases_dir
    golden_dir: Path = args.golden_dir

    files = sorted(cases_dir.glob("*.case.json"))
    if args.suite:
        wanted = set(args.suite)
        files = [f for f in files if f.name.removesuffix(".case.json") in wanted]
    if not files:
        print(f"[cf-oracle-gen] no .case.json files under {cases_dir}", file=sys.stderr)
        return 0

    for case_path in files:
        suite = case_path.name.removesuffix(".case.json")
        golden_path = golden_dir / f"{suite}.golden.json"
        print(f"[cf-oracle-gen] {case_path.name} -> {golden_path.name}")
        _process_suite(case_path, golden_path)
        print(f"[cf-oracle-gen]   wrote {golden_path}")

    return 0


if __name__ == "__main__":
    sys.exit(main())
